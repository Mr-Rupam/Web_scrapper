import csv
from serpapi import GoogleSearch
from openpyxl import Workbook
from openpyxl.styles import Font
import time

# Your SerpAPI key
SERPAPI_API_KEY = "PASTE_API_KEY_HERE"  # Replace with your key

# Define the search roles
search_terms = ["Marketing Head", "Branding Head"]

# Function to search using SerpAPI
def search_linkedin_profiles(company_name, role, api_key):
    query = f"{company_name} {role} site:linkedin.com/in/"
    search = GoogleSearch({
        "q": query,
        "api_key": api_key,
        "num": 3,
        "hl": "en",
        "gl": "in"
    })

    try:
        results = search.get_dict()
    except Exception as e:
        print(f"Error fetching results for {query}: {e}")
        return []

    linkedin_profiles = []
    for result in results.get("organic_results", []):
        link = result.get("link", "")
        title = result.get("title", "")
        if "linkedin.com/in/" in link:
            linkedin_profiles.append((title, role, link))

    return linkedin_profiles

# Read companies from CSV
input_file = "input_companies.csv"
with open(input_file, "r") as csvfile:
    companies = [row[0].strip() for row in csv.reader(csvfile) if row]
    total_companies = len(companies)

# Prepare Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "LinkedIn Profiles"
ws.append(["Profile Name", "Company Name", "Role", "LinkedIn URL"])

# Loop through companies and search
for i, company in enumerate(companies):
    print(f"Searching for profiles in: {company}")
    for term in search_terms:
        profiles = search_linkedin_profiles(company, term, SERPAPI_API_KEY)
        for profile in profiles:
            name, role, url = profile
            cell = ws.cell(row=ws.max_row + 1, column=4)
            cell.value = "Link"
            cell.hyperlink = url
            cell.font = Font(color="0000FF", underline="single")
            ws.cell(row=ws.max_row, column=1).value = name
            ws.cell(row=ws.max_row, column=2).value = company
            ws.cell(row=ws.max_row, column=3).value = role

    # Optional: Simple progress indicator
    progress = int((i + 1) / total_companies * 40)
    print(f"[{('#' * progress)}{(' ' * (40 - progress))}] {i+1}/{total_companies}", end="\r")
    time.sleep(0.5)

# Save to Excel
output_file = "linkedin_profiles.xlsx"
wb.save(output_file)
print(f"\n✅ Done! LinkedIn profiles saved to '{output_file}'")
